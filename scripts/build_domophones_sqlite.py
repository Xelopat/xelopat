from __future__ import annotations

import argparse
import re
import sqlite3
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from normalize_domophones import (
    DEFAULT_SOURCE,
    WordReader,
    clean_text,
    clean_title,
    discover_files,
    first_non_empty,
    location_from_path,
    parse_index_rows,
    pythoncom,
    rel_path,
    zip_mimetype,
)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "data" / "domophones.sqlite"
DEFAULT_SUPPLEMENT = DEFAULT_SOURCE / "Доп База.xlsx"
MOSCOW = "Москва"
DEFAULT_FILE_PROGRESS = 30
DEFAULT_SUPPLEMENT_PROGRESS = 10000

STREET_TYPE_ALIASES = {
    "ул": "ул.",
    "улица": "ул.",
    "пр": "пр.",
    "пр-д": "пр.",
    "проезд": "пр.",
    "пер": "пер.",
    "переулок": "пер.",
    "просп": "проспект",
    "пр-т": "проспект",
    "проспект": "проспект",
    "б-р": "бульвар",
    "бул": "бульвар",
    "бульвар": "бульвар",
    "ш": "шоссе",
    "шоссе": "шоссе",
    "аллея": "аллея",
    "пл": "пл.",
    "площадь": "пл.",
    "наб": "наб.",
    "набережная": "наб.",
    "туп": "туп.",
    "тупик": "туп.",
}

HOUSE_LINE_RE = re.compile(
    r"^\d+[а-яА-Яa-zA-Z]?(?:[/.-]\d+[а-яА-Яa-zA-Z]?)?(?:\s*к\.?\s*\d+[а-яА-Яa-zA-Z]?)?$",
    re.I,
)
ENTRANCE_LINE_RE = re.compile(r"^(\d{1,3})\)\s*(.*)$")
CODE_LIKE_RE = re.compile(r"(?iu)(?:\d+\s*)?(?:реш|[#*]|[вкдdз])\s*\d+|\d{3,8}")


class DebugLog:
    def __init__(self, enabled: bool = False, path: Path | None = None) -> None:
        self.enabled = enabled
        self.path = path
        self.started_at = time.perf_counter()
        if self.enabled and self.path:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self.path.write_text("", encoding="utf-8")

    def write(self, message: str, always: bool = False) -> None:
        if not self.enabled and not always:
            return
        elapsed = time.perf_counter() - self.started_at
        line = f"[{elapsed:9.2f}s] {message}"
        print(line, file=sys.stderr, flush=True)
        if self.path:
            with self.path.open("a", encoding="utf-8") as handle:
                handle.write(line + "\n")


def table_count(conn: sqlite3.Connection, table: str) -> int:
    return one(conn, f"SELECT COUNT(*) FROM {table}", ())


def connect_db(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = NORMAL")
    conn.execute("PRAGMA foreign_keys = ON")
    create_schema(conn)
    return conn


def remove_database(path: Path) -> None:
    for candidate in (path, path.with_name(path.name + "-wal"), path.with_name(path.name + "-shm")):
        if candidate.exists():
            candidate.unlink()


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS cities (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS streets (
            id INTEGER PRIMARY KEY,
            city_id INTEGER NOT NULL REFERENCES cities(id),
            name TEXT NOT NULL,
            UNIQUE(city_id, name)
        );

        CREATE TABLE IF NOT EXISTS houses (
            id INTEGER PRIMARY KEY,
            street_id INTEGER NOT NULL REFERENCES streets(id),
            house_number TEXT NOT NULL,
            building TEXT NOT NULL DEFAULT '',
            raw_house TEXT NOT NULL,
            UNIQUE(street_id, house_number, building)
        );

        CREATE TABLE IF NOT EXISTS entrances (
            id INTEGER PRIMARY KEY,
            house_id INTEGER NOT NULL REFERENCES houses(id),
            entrance_number TEXT NOT NULL DEFAULT '',
            UNIQUE(house_id, entrance_number)
        );

        CREATE TABLE IF NOT EXISTS sources (
            id INTEGER PRIMARY KEY,
            path TEXT NOT NULL UNIQUE,
            source_type TEXT NOT NULL,
            district TEXT NOT NULL DEFAULT '',
            area TEXT NOT NULL DEFAULT '',
            subarea TEXT NOT NULL DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS codes (
            id INTEGER PRIMARY KEY,
            entrance_id INTEGER NOT NULL REFERENCES entrances(id),
            code TEXT NOT NULL,
            raw TEXT NOT NULL DEFAULT '',
            source_id INTEGER REFERENCES sources(id),
            source_row INTEGER,
            UNIQUE(entrance_id, code, source_id, source_row)
        );

        CREATE TABLE IF NOT EXISTS index_rows (
            id INTEGER PRIMARY KEY,
            source_id INTEGER NOT NULL REFERENCES sources(id),
            sheet TEXT NOT NULL DEFAULT '',
            street TEXT NOT NULL DEFAULT '',
            date_text TEXT NOT NULL DEFAULT '',
            raw TEXT NOT NULL DEFAULT '',
            source_row INTEGER
        );

        CREATE INDEX IF NOT EXISTS idx_streets_name ON streets(name);
        CREATE INDEX IF NOT EXISTS idx_houses_lookup ON houses(street_id, house_number, building);
        CREATE INDEX IF NOT EXISTS idx_codes_code ON codes(code);
        """
    )


def one(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...]) -> int:
    cur = conn.execute(sql, params)
    row = cur.fetchone()
    if row is None:
        raise RuntimeError(f"Expected row for query: {sql}")
    return int(row[0])


def get_city(conn: sqlite3.Connection, name: str = MOSCOW) -> int:
    conn.execute("INSERT OR IGNORE INTO cities(name) VALUES (?)", (name,))
    return one(conn, "SELECT id FROM cities WHERE name = ?", (name,))


def get_street(conn: sqlite3.Connection, city_id: int, name: str) -> int:
    name = normalize_street(name)
    conn.execute("INSERT OR IGNORE INTO streets(city_id, name) VALUES (?, ?)", (city_id, name))
    return one(conn, "SELECT id FROM streets WHERE city_id = ? AND name = ?", (city_id, name))


def get_house(conn: sqlite3.Connection, street_id: int, raw_house: str) -> int:
    house_number, building = split_house(raw_house)
    conn.execute(
        """
        INSERT OR IGNORE INTO houses(street_id, house_number, building, raw_house)
        VALUES (?, ?, ?, ?)
        """,
        (street_id, house_number, building, clean_title(raw_house)),
    )
    return one(
        conn,
        "SELECT id FROM houses WHERE street_id = ? AND house_number = ? AND building = ?",
        (street_id, house_number, building),
    )


def get_entrance(conn: sqlite3.Connection, house_id: int, number: Any) -> int:
    entrance_number = "" if number is None else clean_text(number)
    conn.execute(
        "INSERT OR IGNORE INTO entrances(house_id, entrance_number) VALUES (?, ?)",
        (house_id, entrance_number),
    )
    return one(
        conn,
        "SELECT id FROM entrances WHERE house_id = ? AND entrance_number = ?",
        (house_id, entrance_number),
    )


def get_source(
    conn: sqlite3.Connection,
    path: str,
    source_type: str,
    district: str = "",
    area: str = "",
    subarea: list[str] | str | None = None,
) -> int:
    subarea_text = "/".join(subarea) if isinstance(subarea, list) else (subarea or "")
    conn.execute(
        """
        INSERT OR IGNORE INTO sources(path, source_type, district, area, subarea)
        VALUES (?, ?, ?, ?, ?)
        """,
        (path, source_type, district, area, subarea_text),
    )
    return one(conn, "SELECT id FROM sources WHERE path = ?", (path,))


def add_code(
    conn: sqlite3.Connection,
    entrance_id: int,
    code: str,
    raw: str,
    source_id: int,
    source_row: int | None,
) -> None:
    code = normalize_code(code)
    if not code:
        return
    conn.execute(
        """
        INSERT OR IGNORE INTO codes(entrance_id, code, raw, source_id, source_row)
        VALUES (?, ?, ?, ?, ?)
        """,
        (entrance_id, code, clean_text(raw), source_id, source_row),
    )


def normalize_street(value: str) -> str:
    text = clean_title(value)
    text = re.sub(r"^г\.?\s*москва\s*,?\s*", "", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip(" ,.")

    match = re.match(r"^([А-Яа-яA-Za-z.-]+)\.?\s+(.+)$", text)
    if match:
        first = match.group(1).lower().rstrip(".")
        rest = match.group(2).strip(" ,.")
        if first in STREET_TYPE_ALIASES:
            return f"{rest} {STREET_TYPE_ALIASES[first]}".strip()

    return text


def split_house(value: str) -> tuple[str, str]:
    text = clean_title(value).lower()
    text = re.sub(r"^(д\.?|дом)\s*", "", text).strip()
    text = text.replace("корп.", "к").replace("корпус", "к")
    text = re.sub(r"\s+", "", text)

    match = re.match(r"^(.+?)(?:к\.?([0-9а-яa-z/-]+))$", text, flags=re.I)
    if match:
        return match.group(1), match.group(2)
    return text, ""


def normalize_code(value: str) -> str:
    text = clean_text(value).lower()
    text = text.replace("реш", "#")
    text = re.sub(r"\s+", "", text)
    return text.strip(" ,.;")


def parse_supplement_address(address: str) -> tuple[str, str]:
    text = clean_text(address)
    text = re.sub(r"^г\.?\s*москва\s*,\s*", "", text, flags=re.I)
    parts = [part.strip() for part in text.split(",") if part.strip()]
    street = parts[0] if parts else text
    house_parts = []
    for part in parts[1:]:
        if re.search(r"^(д|дом|к|корпус|корп)\.?\s*", part, flags=re.I):
            house_parts.append(part)
    house = " ".join(house_parts)
    house = re.sub(r"\s*,?\s*к\.?\s*", "к", house, flags=re.I)
    house = re.sub(r"^(д\.?|дом)\s*", "", house, flags=re.I).strip()
    return normalize_street(street), house


def split_supplement_codes(values: list[Any]) -> list[str]:
    text = "\n".join(clean_text(value) for value in values if clean_text(value))
    parts = re.split(r"[,;\n]+", text)
    return [clean_text(part) for part in parts if clean_text(part)]


def word_blocked_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return "заблокирован параметрами блокировки файлов" in text or "file block" in text


def read_blocked_doc(path: Path, source_dir: Path) -> list[dict[str, Any]]:
    lines = extract_legacy_doc_lines(path)
    if not lines:
        return []

    street_index = find_street_line(lines, path)
    if street_index is None:
        return []

    street = clean_title(lines[street_index])
    location = location_from_path(path, source_dir)
    records_by_house: dict[str, dict[str, Any]] = {}
    current_house = ""

    for source_row, line in enumerate(lines[street_index + 1 :], start=1):
        if is_house_line(line):
            current_house = clean_title(line)
            records_by_house.setdefault(
                current_house,
                {
                    **location,
                    "street": street,
                    "house": current_house,
                    "entrances": [],
                    "raw_codes": "",
                    "source_file": rel_path(path, source_dir),
                    "source_table": 0,
                    "source_row": source_row,
                },
            )
            continue

        if not current_house or is_noise_line(line):
            continue

        parsed_line = parse_legacy_code_line(line)
        if parsed_line is None:
            continue

        record = records_by_house[current_house]
        record["entrances"].append(parsed_line)
        record["raw_codes"] = (record["raw_codes"] + "\n" + parsed_line["raw"]).strip()

    return [record for record in records_by_house.values() if record["entrances"]]


def extract_legacy_doc_lines(path: Path) -> list[str]:
    data = path.read_bytes()
    decoded = data.decode("cp1251", errors="ignore")
    allowed = set("#*()/.,-– ")
    chars: list[str] = []
    for char in decoded:
        if char in "\r\n\t\x07":
            chars.append("\n")
        elif char.isalnum() or char in allowed:
            chars.append(char)
        else:
            chars.append("\n")

    lines: list[str] = []
    for raw_line in "".join(chars).splitlines():
        line = clean_text(raw_line)
        if not line or len(line) > 120:
            continue
        if line.lower() in {"root entry", "worddocument", "summaryinformation", "documentsummaryinformation"}:
            continue
        if has_cyrillic_or_digit(line):
            lines.append(line)
    return lines


def find_street_line(lines: list[str], path: Path) -> int | None:
    stem_words = [word.lower() for word in re.findall(r"[а-яА-Яa-zA-Z0-9]+", path.stem) if len(word) > 2]
    for index, line in enumerate(lines):
        lowered = line.lower()
        if any(word in lowered for word in stem_words) and has_cyrillic_or_digit(line):
            return index

    for index, line in enumerate(lines):
        lowered = line.lower()
        if any(marker in lowered for marker in (" ул", "улица", "шоссе", "просп", "бульвар", "проезд", "пер")):
            return index

    return None


def has_cyrillic_or_digit(value: str) -> bool:
    return any("а" <= char.lower() <= "я" or char.isdigit() for char in value)


def is_house_line(value: str) -> bool:
    line = clean_text(value)
    if not line or ")" in line:
        return False
    return bool(HOUSE_LINE_RE.match(line))


def is_noise_line(value: str) -> bool:
    line = clean_text(value).lower()
    if not line:
        return True
    if line.startswith("times new roman") or line.startswith("arial"):
        return True
    if line in {"ктв", "инет", "(ктв)", "(инет)", "конс"}:
        return True
    return False


def parse_legacy_code_line(line: str) -> dict[str, Any] | None:
    cleaned = clean_text(line)
    match = ENTRANCE_LINE_RE.match(cleaned)
    entrance: int | None = None
    code_text = cleaned
    if match:
        entrance = int(match.group(1))
        code_text = clean_text(match.group(2))

    tokens = [clean_text(match.group(0)) for match in CODE_LIKE_RE.finditer(code_text)]
    if not tokens and not re.search(r"\d", code_text):
        return None

    return {
        "entrance": entrance,
        "raw": cleaned,
        "code": code_text,
        "tokens": tokens,
    }


def read_spreadsheet_index(path: Path, source_dir: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    mimetype = zip_mimetype(path)

    if mimetype == "application/vnd.oasis.opendocument.spreadsheet" or suffix == ".ods":
        sheets = read_ods_index(path)
    elif suffix == ".xlsx":
        sheets = read_xlsx_index(path)
    elif suffix in {".xls", ".xlt"}:
        sheets = read_xls_index(path)
    else:
        sheets = []

    return {
        **location_from_path(path, source_dir),
        "source_file": rel_path(path, source_dir),
        "sheets": sheets,
    }


def read_xlsx_index(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            rows = [
                [clean_text(value) for value in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            rows = [row for row in rows if any(row)]
            sheets.append({"name": clean_text(sheet.title), "title": first_non_empty(rows), "rows": parse_index_rows(rows)})
        return sheets
    finally:
        workbook.close()


def read_xls_index(path: Path) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("xlrd is required to read legacy .xls/.xlt index files") from exc

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    sheets: list[dict[str, Any]] = []
    try:
        for sheet in workbook.sheets():
            rows: list[list[str]] = []
            for row_index in range(sheet.nrows):
                row = [format_xlrd_cell(sheet.cell(row_index, col_index), workbook) for col_index in range(sheet.ncols)]
                if any(row):
                    rows.append(row)
            sheets.append({"name": clean_text(sheet.name), "title": first_non_empty(rows), "rows": parse_index_rows(rows)})
    finally:
        workbook.release_resources()
    return sheets


def format_xlrd_cell(cell: Any, workbook: Any) -> str:
    try:
        import xlrd
    except ImportError:
        return clean_text(cell.value)

    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            parts = xlrd.xldate_as_tuple(cell.value, workbook.datemode)
            if parts[:3] == (0, 0, 0):
                return clean_text(cell.value)
            return f"{parts[2]:02d}.{parts[1]:02d}.{parts[0]:04d}"
        except Exception:
            return clean_text(cell.value)
    if cell.ctype == xlrd.XL_CELL_NUMBER and float(cell.value).is_integer():
        return str(int(cell.value))
    return clean_text(cell.value)


def read_ods_index(path: Path) -> list[dict[str, Any]]:
    ns = {
        "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
        "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
        "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
    }

    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("content.xml"))

    tables = root.findall(".//table:table", ns)
    sheets: list[dict[str, Any]] = []
    for table in tables:
        sheet_name = clean_text(table.attrib.get(f"{{{ns['table']}}}name", ""))
        rows: list[list[str]] = []
        for row in table.findall("table:table-row", ns):
            repeat_rows = int(row.attrib.get(f"{{{ns['table']}}}number-rows-repeated", "1"))
            parsed_row: list[str] = []
            for cell in row.findall("table:table-cell", ns):
                repeat_cols = int(cell.attrib.get(f"{{{ns['table']}}}number-columns-repeated", "1"))
                value = ods_cell_text(cell, ns)
                # Repeated empty cells can describe the rest of a huge sheet; keeping a few is enough for indexes.
                max_repeat = repeat_cols if value else min(repeat_cols, 8)
                parsed_row.extend([value] * max_repeat)
            parsed_row = trim_empty_tail(parsed_row)
            if any(parsed_row):
                rows.extend([parsed_row] * min(repeat_rows, 1))
        sheets.append({"name": sheet_name, "title": first_non_empty(rows), "rows": parse_index_rows(rows)})
    return sheets


def ods_cell_text(cell: ET.Element, ns: dict[str, str]) -> str:
    texts: list[str] = []
    for paragraph in cell.findall(".//text:p", ns):
        value = "".join(paragraph.itertext())
        if value:
            texts.append(value)
    if texts:
        return clean_text("\n".join(texts))
    return clean_text(cell.attrib.get(f"{{{ns['office']}}}value", ""))


def trim_empty_tail(row: list[str]) -> list[str]:
    while row and not row[-1]:
        row.pop()
    return row


def import_address_record(conn: sqlite3.Connection, city_id: int, record: dict[str, Any]) -> None:
    source_id = get_source(
        conn,
        record["source_file"],
        "office_document",
        record.get("district", ""),
        record.get("area", ""),
        record.get("subarea", []),
    )
    street_id = get_street(conn, city_id, record["street"])
    house_id = get_house(conn, street_id, record["house"])

    for entrance in record.get("entrances", []):
        entrance_id = get_entrance(conn, house_id, entrance.get("entrance"))
        tokens = entrance.get("tokens") or []
        if not tokens and re.search(r"\d", entrance.get("code", "")):
            tokens = [entrance.get("code", "")]
        for token in tokens:
            add_code(conn, entrance_id, token, entrance.get("raw", ""), source_id, record.get("source_row"))


def import_index(conn: sqlite3.Connection, item: dict[str, Any]) -> None:
    source_id = get_source(
        conn,
        item["source_file"],
        "street_index",
        item.get("district", ""),
        item.get("area", ""),
        item.get("subarea", []),
    )
    for sheet in item.get("sheets", []):
        for row in sheet.get("rows", []):
            conn.execute(
                """
                INSERT INTO index_rows(source_id, sheet, street, date_text, raw, source_row)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    source_id,
                    sheet.get("name", ""),
                    normalize_street(row.get("street", "")),
                    clean_text(row.get("date", "")),
                    " | ".join(row.get("raw", [])),
                    row.get("source_row"),
                ),
            )


def import_office_sources(
    conn: sqlite3.Connection,
    source_dir: Path,
    offset: int,
    limit: int | None,
    progress: int,
    debug: DebugLog,
) -> None:
    files = [
        path
        for path in discover_files(source_dir)
        if path.name.lower() != DEFAULT_SUPPLEMENT.name.lower()
    ]
    if offset:
        files = files[offset:]
    if limit is not None:
        files = files[:limit]

    debug.write(
        f"Office import planned: files={len(files)}, offset={offset}, limit={limit if limit is not None else 'all'}",
        always=True,
    )
    city_id = get_city(conn)
    debug.write("Starting COM: Word")
    pythoncom.CoInitialize()
    word = WordReader()
    debug.write("COM ready: Word")
    try:
        for file_index, path in enumerate(files, start=1):
            relative = rel_path(path, source_dir)
            if progress and (file_index == 1 or file_index % progress == 0):
                debug.write(f"Office progress {file_index}/{len(files)}: {relative}", always=True)
            suffix = path.suffix.lower()
            file_started = time.perf_counter()
            before_codes = table_count(conn, "codes")
            before_index_rows = table_count(conn, "index_rows")
            before_sources = table_count(conn, "sources")
            debug.write(f"START office {file_index}/{len(files)} {suffix}: {relative}")
            try:
                if suffix in {".doc", ".docx", ".odt"}:
                    try:
                        records, _notes = word.read(path, source_dir)
                    except Exception as exc:
                        if suffix == ".doc" and word_blocked_error(exc):
                            records = read_blocked_doc(path, source_dir)
                            debug.write(f"READ blocked legacy document fallback: records={len(records)} file={relative}", always=True)
                        else:
                            raise
                    debug.write(f"READ office document: records={len(records)} file={relative}")
                    for record in records:
                        import_address_record(conn, city_id, record)
                elif suffix in {".xls", ".xlsx", ".ods", ".xlt"}:
                    # The old spreadsheet files are street/date indexes, not code rows.
                    # Read them without Excel COM; some files have a wrong extension and can hang Excel.
                    index = read_spreadsheet_index(path, source_dir)
                    debug.write(
                        "READ office spreadsheet: "
                        f"sheets={len(index.get('sheets', []))} "
                        f"rows={sum(len(sheet.get('rows', [])) for sheet in index.get('sheets', []))} "
                        f"file={relative}"
                    )
                    import_index(conn, index)
                conn.commit()
                elapsed = time.perf_counter() - file_started
                debug.write(
                    "DONE office "
                    f"{file_index}/{len(files)} in {elapsed:.2f}s: {relative}; "
                    f"+codes={table_count(conn, 'codes') - before_codes}, "
                    f"+index_rows={table_count(conn, 'index_rows') - before_index_rows}, "
                    f"+sources={table_count(conn, 'sources') - before_sources}"
                )
            except Exception as exc:
                location = location_from_path(path, source_dir)
                get_source(
                    conn,
                    rel_path(path, source_dir),
                    f"error: {exc}",
                    location.get("district", ""),
                    location.get("area", ""),
                    location.get("subarea", []),
                )
                conn.commit()
                elapsed = time.perf_counter() - file_started
                debug.write(f"ERROR office {file_index}/{len(files)} in {elapsed:.2f}s: {relative}: {exc}", always=True)
    finally:
        debug.write("Closing COM: Word")
        word.close()
        pythoncom.CoUninitialize()
        debug.write("COM closed")


def import_supplement(conn: sqlite3.Connection, source_dir: Path, path: Path, progress: int, debug: DebugLog) -> None:
    if not path.exists():
        debug.write(f"Supplement not found: {path}", always=True)
        return

    city_id = get_city(conn)
    source_id = get_source(conn, rel_path(path, source_dir), "supplement_xlsx")
    debug.write(f"START supplement: {path}", always=True)
    started = time.perf_counter()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            debug.write(f"Supplement sheet: {sheet.title}, max_row={sheet.max_row}, max_column={sheet.max_column}", always=True)
            before_codes = table_count(conn, "codes")
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_index == 1:
                    continue
                if progress and row_index % progress == 0:
                    debug.write(
                        f"Supplement progress row={row_index}/{sheet.max_row} sheet={sheet.title}; "
                        f"+codes={table_count(conn, 'codes') - before_codes}",
                        always=True,
                    )

                address = clean_text(row[0] if len(row) > 0 else "")
                entrance = clean_text(row[1] if len(row) > 1 else "")
                if not address:
                    continue

                street, house = parse_supplement_address(address)
                if not street or not house:
                    continue

                street_id = get_street(conn, city_id, street)
                house_id = get_house(conn, street_id, house)
                entrance_id = get_entrance(conn, house_id, entrance)
                for code in split_supplement_codes(list(row[2:])):
                    add_code(conn, entrance_id, code, code, source_id, row_index)

                if row_index % 1000 == 0:
                    conn.commit()
    finally:
        workbook.close()
    conn.commit()
    debug.write(
        f"DONE supplement in {time.perf_counter() - started:.2f}s: {path}",
        always=True,
    )


def print_stats(conn: sqlite3.Connection, output: Path) -> None:
    stats = {
        "streets": one(conn, "SELECT COUNT(*) FROM streets", ()),
        "houses": one(conn, "SELECT COUNT(*) FROM houses", ()),
        "entrances": one(conn, "SELECT COUNT(*) FROM entrances", ()),
        "codes": one(conn, "SELECT COUNT(*) FROM codes", ()),
        "sources": one(conn, "SELECT COUNT(*) FROM sources", ()),
        "index_rows": one(conn, "SELECT COUNT(*) FROM index_rows", ()),
    }
    print(f"SQLite ready: {output}")
    print(", ".join(f"{key}={value}" for key, value in stats.items()))


def main() -> int:
    parser = argparse.ArgumentParser(description="Build a normalized SQLite database from intercom source files.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--supplement", type=Path, default=DEFAULT_SUPPLEMENT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--offset", type=int, default=0, help="Skip the first N old Office source files.")
    parser.add_argument("--limit", type=int, default=None, help="Process only N old Office source files.")
    parser.add_argument("--skip-office", action="store_true", help="Import only the supplemental XLSX file.")
    parser.add_argument("--skip-supplement", action="store_true", help="Import only old Office source files.")
    parser.add_argument("--progress", type=int, default=DEFAULT_FILE_PROGRESS, help="Print old Office progress every N files.")
    parser.add_argument(
        "--supplement-progress",
        type=int,
        default=DEFAULT_SUPPLEMENT_PROGRESS,
        help="Print supplemental XLSX progress every N rows.",
    )
    parser.add_argument("--debug", action="store_true", help="Print detailed timings for every source file.")
    parser.add_argument("--log", type=Path, default=ROOT / "data" / "domophones_import.log")
    parser.set_defaults(reset=True)
    parser.add_argument("--reset", dest="reset", action="store_true", help="Delete the output database before importing.")
    parser.add_argument("--no-reset", dest="reset", action="store_false", help="Append to the existing database.")
    args = parser.parse_args()

    source_dir = args.source.resolve()
    output = args.output.resolve()
    supplement = args.supplement.resolve()
    if args.reset:
        remove_database(output)

    debug = DebugLog(args.debug, args.log.resolve() if args.debug else None)
    debug.write(f"Output database: {output}", always=True)
    debug.write(f"Source directory: {source_dir}", always=True)
    debug.write(f"Supplement file: {supplement}", always=True)

    conn = connect_db(output)
    try:
        if not args.skip_office:
            import_office_sources(conn, source_dir, args.offset, args.limit, args.progress, debug)
        if not args.skip_supplement:
            import_supplement(conn, source_dir, supplement, args.supplement_progress, debug)
        print_stats(conn, output)
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
